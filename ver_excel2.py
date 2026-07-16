import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
excel = 'C:/AGF_Coleta/Inventario_Florestal/Planilha padrão  - ÚNICA.xlsx'
wb = openpyxl.load_workbook(excel)

# Ler as abas de caracterização com mais detalhes
for sheet_name in ['Caracterização FESD', 'Caracterização Cerrado', 'Caracterização CR', 'Floristica caminhamento CR']:
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} ===")
    
    # Ler todas as linhas até 5
    for row in range(1, min(ws.max_row + 1, 6)):
        data = []
        for col in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row=row, column=col)
            data.append(str(cell.value) if cell.value else "")
        print(f"Linha {row}: {data}")
