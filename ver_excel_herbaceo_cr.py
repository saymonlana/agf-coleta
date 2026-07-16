import sys
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
excel = 'C:/AGF_Coleta/Inventario_Florestal/Planilha padrão  - ÚNICA.xlsx'
wb = openpyxl.load_workbook(excel)

ws = wb['Parcela (herbáceo CR)']
print("=== Parcela (herbáceo CR) - Excel ===")
print(f"Linhas: {ws.max_row}, Colunas: {ws.max_column}")

# Ler linha 5 (que parece ter os nomes dos campos)
for row in range(1, min(ws.max_row + 1, 8)):
    data = []
    for col in range(1, min(ws.max_column + 1, 25)):
        cell = ws.cell(row=row, column=col)
        data.append(str(cell.value) if cell.value else "")
    print(f"Linha {row}: {data}")
