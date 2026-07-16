import sys
sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
    print("openpyxl disponível!")
    
    excel = 'C:/AGF_Coleta/Inventario_Florestal/Planilha padrão  - ÚNICA.xlsx'
    wb = openpyxl.load_workbook(excel)
    
    print(f"\nAbas encontradas: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n=== {sheet_name} ===")
        print(f"Linhas: {ws.max_row}, Colunas: {ws.max_column}")
        
        # Mostrar primeira linha (cabeçalhos)
        if ws.max_row > 0:
            headers = []
            for col in range(1, min(ws.max_column + 1, 30)):
                cell = ws.cell(row=1, column=col)
                headers.append(str(cell.value) if cell.value else "")
            print(f"Cabeçalhos (Linha 1): {headers}")
        
        # Mostrar segunda linha se existir (possíveis aliases)
        if ws.max_row > 1:
            row2 = []
            for col in range(1, min(ws.max_column + 1, 30)):
                cell = ws.cell(row=2, column=col)
                row2.append(str(cell.value) if cell.value else "")
            print(f"Linha 2: {row2}")
        
        print()
        
except ImportError:
    print("openpyxl não instalado")
except Exception as e:
    print(f"Erro: {e}")
